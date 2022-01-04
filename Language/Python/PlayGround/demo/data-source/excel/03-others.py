import openpyxl
from openpyxl.styles import Font
import random

# 1. 设置斜体与字号
wb = openpyxl.Workbook()
sheet = wb['Sheet']
italic24Font = Font(size=24, italic=True)
sheet['A1'].font = italic24Font
sheet['A1'] = "Hello World."
wb.save('data/styles.xlsx')

# 2. 设置字体...
wb = openpyxl.Workbook()
sheet = wb['Sheet']

fontObj1 = Font(name='Times New Roman', bold=True)
sheet['A1'].font = fontObj1
sheet['A1'] = 'Bold Times New Roman'

fontObj2 = Font(size=24, italic=True)
sheet['B3'].font = fontObj2
sheet['B3'] = '24 pt Italic'

wb.save('data/styles2.xlsx')

# 3. 公式
wb = openpyxl.Workbook()
sheet = wb.active
testdata = [random.randint(10, 100) for i in range(8)]
for i, data in enumerate(testdata):
    sheet.cell(row=i+1, column=2, value=data)
sheet['A9'] = 'Total:'
sheet['B9'] = '=SUM(B1:B8)'
wb.save('data/formula.xlsx')

# 4. 调整行高与列宽
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('data/dimensions.xlsx')

# 5. 合并拆分单元格
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two metged cells.'
wb.save('data/merged.xlsx')

wb = openpyxl.load_workbook('data/merged.xlsx')
sheet = wb.active
sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')
wb.save('data/splitd.xlsx')

# 6. 冻结解冻窗格
wb = openpyxl.load_workbook('data/produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2'  # 参数可以为Cell对象或单元格坐标字符串，值为None或'A1'为解冻操作
wb.save('data/freezed.xlsx')

# 7. 图表
"""
创建图表的步骤：
    1．从一个矩形区域选择单元格来创建一个Reference对象
    2．通过传入Reference对象来创建一个Series对象
    3．创建一个Chart对象
    4．将Series对象添加到Chart对象
    5．可选地设置Chart对象的drawing.top、drawing.left、drawing.width和drawing.height属性
    6．将Chart对象添加到Worksheet对象
Reference对象需要一些解释，Reference对象是通过调用openpyxl.charts.Reference()函数并传入以下3个参数创建的：
    1．包含图表数据的Worksheet对象
    2．两个整数的元组，代表矩形选择区域的左上角单元格，该区域包含图表数据：
        元组中第一个整数是行，第二个整数是列。请注意第一行是1，不是0。
    3．两个整数的元组，代表矩形选择区域的右下角单元格，该区域包含图表数据：
        元组中第一个整数是行，第二个整数是列。
"""

wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11):
    sheet['A'+str(i)] = i

refObj = openpyxl.chart.Reference(
    sheet, min_col=1, min_row=1, max_col=1, max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title='First series')

chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)

sheet.add_chart(chartObj, 'C5')
wb.save('data/sampleChart.xlsx')
