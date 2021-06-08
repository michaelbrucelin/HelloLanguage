# pip install openpyxl

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# 1. 获取工作簿
wb = openpyxl.load_workbook('data/example.xlsx')
type(wb)       # <class 'openpyxl.workbook.workbook.Workbook'>

# 2. 获取工作表
wb = openpyxl.load_workbook('data/example.xlsx')
wb.sheetnames  # ['Sheet1', 'Sheet2', 'Sheet3']
sheet = wb['Sheet3']
sheet          # <Worksheet "Sheet3">
type(sheet)    # <class 'openpyxl.worksheet.worksheet.Worksheet'>
sheet.title    # 'Sheet3'
activeSheet = wb.active
activeSheet    # <Worksheet "Sheet1">

# 3.获取单元格
wb = openpyxl.load_workbook('data/example.xlsx')
sheet = wb['Sheet1']
sheet.max_row     # 7
sheet.max_column  # 3

sheet['A1']                        # <Cell 'Sheet1'.A1>
sheet['A1'].value                  # datetime.datetime(2015, 4, 5, 13, 34, 2)
sheet.cell(row=1, column=2)        # <Cell 'Sheet1'.B1>
sheet.cell(row=1, column=2).value  # 'Apples'

c = sheet['B1']
c.value            # 'Apples'
'Row %s, Column %s is %s' % (c.row, c.column, c.value)
# 'Row 1, Column 2 is Apples'
'Cell %s is %s' % (c.coordinate, c.value)  # 'Cell B1 is Apples'
sheet['C1'].value  # 73

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
# 1 Apples
# 3 Pears
# 5 Apples
# 7 Strawberries

# 3. 列字母与数字之间的转换
# from openpyxl.utils import get_column_letter, column_index_from_string
get_column_letter(1)    # 'A'
get_column_letter(2)    # 'B'
get_column_letter(27)   # 'AA'
get_column_letter(900)  # 'AHP'

wb = openpyxl.load_workbook('data/example.xlsx')
sheet = wb['Sheet1']
get_column_letter(sheet.max_column)  # 'C'
column_index_from_string('A')        # 1
column_index_from_string('AA')       # 27

# 4. 从表中获取行和列
wb = openpyxl.load_workbook('data/example.xlsx')
sheet = wb['Sheet1']
tuple(sheet['A1':'C3'])
# ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>),
#  (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>),
#  (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')
# A1 2015-04-05 13:34:02
# B1 Apples
# C1 73
# --- END OF ROW ---
# A2 2015-04-05 03:41:23
# B2 Cherries
# C2 85
# --- END OF ROW ---
# A3 2015-04-06 12:46:51
# B3 Pears
# C3 14
# --- END OF ROW ---

sheet = wb.active
list(sheet['A1':'C3'])
# [(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>),
#  (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>),
#  (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>)]
list(sheet.rows)
# [(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>),
#  (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>),
#  (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>),
#  (<Cell 'Sheet1'.A4>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.C4>),
#  (<Cell 'Sheet1'.A5>, <Cell 'Sheet1'.B5>, <Cell 'Sheet1'.C5>),
#  (<Cell 'Sheet1'.A6>, <Cell 'Sheet1'.B6>, <Cell 'Sheet1'.C6>),
#  (<Cell 'Sheet1'.A7>, <Cell 'Sheet1'.B7>, <Cell 'Sheet1'.C7>)]
list(sheet.columns)
# [(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>, <Cell 'Sheet1'.A4>, <Cell 'Sheet1'.A5>, <Cell 'Sheet1'.A6>, <Cell 'Sheet1'.A7>),
#  (<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.B5>, <Cell 'Sheet1'.B6>, <Cell 'Sheet1'.B7 >),
#  (<Cell 'Sheet1'.C1>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.C3>, <Cell 'Sheet1'.C4>, <Cell 'Sheet1'.C5>, <Cell 'Sheet1'.C6>, <Cell 'Sheet1'.C7 >)]

for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)
# Apples
# Cherries
# Pears
# Oranges
# Apples
# Bananas
# Strawberries
