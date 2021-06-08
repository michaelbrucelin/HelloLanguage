import openpyxl

# 1. 创建一个workbook
wb = openpyxl.Workbook()
wb.sheetnames  # ['Sheet']
sheet = wb.active
sheet.title    # 'Sheet'
sheet.title = 'hello sheet'
wb.sheetnames  # ['hello sheet']
wb.save('data/example_new.xlsx')

# 2. 创建新的工作表
wb = openpyxl.Workbook()
wb.sheetnames      # ['Sheet']
wb.create_sheet()  # <Worksheet "Sheet1">
wb.sheetnames      # ['Sheet', 'Sheet1']
wb.create_sheet(index=0, title='First Sheet')   # <Worksheet "First Sheet">
wb.sheetnames      # ['First Sheet', 'Sheet', 'Sheet1']
wb.create_sheet(index=2, title='Middle Sheet')  # <Worksheet "Middle Sheet">
wb.sheetnames      # ['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet1']
del wb['Middle Sheet']
del wb['Sheet1']
wb.sheetnames      # ['First Sheet', 'Sheet']
# wb.save('data/example_new.xlsx')

# 3. 将值写入单元格
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = "Hello World."
sheet['A1'].value  # 'Hello World.'
# wb.save('data/example_new.xlsx')
